from openpyxl import load_workbook
import os
from apitest.testDemo.utils import log

logger = log.Log()


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.sheet_obj = load_workbook(self.file_name)[self.sheet_name]

    def get_header(self):
        """
        获取第一行的标题行
        """
        headers = []
        for j in range(1, self.sheet_obj.max_column + 1):
            headers.append(self.sheet_obj.cell(1, j).value)
        return headers

    def get_data(self):

        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        # 最大行
        max_row = sheet.max_row
        test_data = []
        headers = self.get_header()
        for i in range(2, max_row + 1):
            res_data = {}
            for j in range(1, sheet.max_column + 1):
                res_data[headers[j - 1]] = sheet.cell(i, j).value
            test_data.append(res_data)
        return test_data

    @staticmethod
    def write_excel(file_name, sheet_name, i, execute_result, test_result):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value = execute_result
        sheet.cell(i, 8).value = test_result

        wb.save(file_name)


if __name__ == '__main__':
    file_name = 'D:/pyworkpase/apiautos/apitest/testDemo/testData/logindata .xlsx'
    test_data = DoExcel(file_name, 'login').get_data()
    print(test_data[0])
    from apitest.testDemo.login_test.request_tools import RequestApi

    res = RequestApi(url=test_data[0]['http'], data=eval(test_data[0]['data']), method=test_data[0]['method'],
                     headers=eval(test_data[0]['header'])).req()
    DoExcel.write_excel(file_name, 'login', test_data[1]['case_id'], res, 'PASS')
    print(res)
